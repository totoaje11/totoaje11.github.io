import paramiko
import warnings
import sys
import os
import subprocess
import argparse
import logging
import concurrent.futures
import glob
import re
from pathlib import Path
import time
import resource
import shutil
import json
import pandas as pd
import numpy as np
import math
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from collections import Counter
import seaborn as sns
import matplotlib.pyplot as plt
import matplotlib.image as mpimg
import plotly.graph_objects as go
from plotly.subplots import make_subplots

####################################

ref_fasta='/data/home/heewon/ref/HG19/hg19.fa'

#####################################tool docker
fastp = 'quay.io/biocontainers/fastp:0.23.2--h79da9fb_0'
bwa = 'quay.io/biocontainers/bwa:0.7.17--hed695b0_7'
samtools = 'quay.io/biocontainers/samtools:1.16.1--h6899075_1'
bicseq2 = 'mwyczalkowski/bicseq2:latest'
picard = 'quay.io/biocontainers/picard:3.0.0--hdfd78af_1'
##############################################

now = datetime.now()
formatted_time = now.strftime("%y%m%d_%H%M%S")

sample_dict = {}


parser = argparse.ArgumentParser(description='In-house Pipeline')
parser.add_argument('run_id', help='run_id (example: 220723_M70544_0051_000000000-G9C6V)')
parser.add_argument('sample_id', help='sample_id (example: 190966)')
parser.add_argument('-o', dest='output_dir', default=f'/data/analysis/lp_wgs', help='output directory (default: %(default)s)')
parser.add_argument('-f1', dest='raw_fastq1', default=f'', help='output directory (default: %(default)s)')
parser.add_argument('-f2', dest='raw_fastq2', default=f'', help='output directory (default: %(default)s)')
parser.add_argument('-t', dest='threads', type=int, default=12, help='threads (default: %(default)s)')
args = parser.parse_args()

sample_dict["sample_dir"] = f'{args.output_dir}/{args.run_id}/{args.sample_id}'
os.makedirs(sample_dict["sample_dir"], exist_ok=True)

logger = logging.getLogger()
logger.setLevel(logging.DEBUG)
formatter = logging.Formatter('[%(levelname)5s - %(asctime)s] %(message)s',"%Y-%m-%d %H:%M:%S")

stream_handler = logging.StreamHandler(sys.stdout)
stream_handler.setFormatter(formatter)
stream_handler.setLevel(logging.INFO)
logger.addHandler(stream_handler)

file_handler = logging.FileHandler(f'{sample_dict["sample_dir"]}/run_{formatted_time}.log','w')
file_handler.setFormatter(formatter)
file_handler.setLevel(logging.DEBUG)
logger.addHandler(file_handler)

start_time = time.time()

def init_env_check():
    logging.info(f'{"-"*20}{"germline variant call In-house Pipeline":^50}{"-"*20}')

    logging.info(f'{"cmd":<15}: {" ".join(sys.argv)}')
    logging.info(f'{"log file":<15}: {sample_dict["sample_dir"]}/preprocess.log')

    logger = logging.getLogger()

    #fastq 파일이 존재하는지 확인
    if args.raw_fastq1 == '' and args.raw_fastq2 == '' :
        sample_dict["raw_fastq1"] = glob.glob(f'/data/raw_data/{args.run_id}/{args.sample_id}*R1*.fastq.gz')[0]
        sample_dict["raw_fastq2"] = glob.glob(f'/data/raw_data/{args.run_id}/{args.sample_id}*R2*.fastq.gz')[0]                
    else :
        sample_dict["raw_fastq1"] = args.raw_fastq1
        sample_dict["raw_fastq2"] = args.raw_fastq2

    if sample_dict["raw_fastq1"] == '' or sample_dict["raw_fastq2"] == '' :
        print('file not exist')
        exit()
    #저장 기간에 따라 폴더를 분리.
    #fastq = 2년, prerocess = 용량이 크므로 1년, result = 영구보관
    sample_dict["sample_fastq_dir"] = f'{args.output_dir}/{args.run_id}/{args.sample_id}/0.fastq'
    sample_dict["sample_preprocess_dir"] = f'{args.output_dir}/{args.run_id}/{args.sample_id}/1.Preprocess'
    sample_dict["sample_bicseq2_dir"] = f'{args.output_dir}/{args.run_id}/{args.sample_id}/2.BICseq2'
    sample_dict["sample_result_dir"] = f'{args.output_dir}/{args.run_id}/{args.sample_id}/3.Result'
    os.makedirs(sample_dict["sample_fastq_dir"], exist_ok=True)
    os.makedirs(sample_dict["sample_preprocess_dir"], exist_ok=True)
    os.makedirs(sample_dict["sample_bicseq2_dir"], exist_ok=True)
    os.makedirs(sample_dict["sample_result_dir"], exist_ok=True)
    shutil.copy(sample_dict["raw_fastq1"], f'{sample_dict["sample_dir"]}/0.fastq/')
    shutil.copy(sample_dict["raw_fastq2"], f'{sample_dict["sample_dir"]}/0.fastq/')
    sample_dict["raw_fastq1"]= f'{sample_dict["sample_dir"]}/0.fastq/{sample_dict["raw_fastq1"].split("/")[-1]}'
    sample_dict["raw_fastq2"]= f'{sample_dict["sample_dir"]}/0.fastq/{sample_dict["raw_fastq2"].split("/")[-1]}'
    

###########################################
######1.preprocess (fastq-> bwa.sorted.bam)
###########################################

######data alignment#####

def run_fastp():
    sample_dict["fastp1"] = f'{sample_dict["sample_fastq_dir"]}/{args.sample_id}_1.fastp.fastq'
    sample_dict["fastp2"] = f'{sample_dict["sample_fastq_dir"]}/{args.sample_id}_2.fastp.fastq'
    cmd = f'docker run -u {os.getuid()}:{os.getgid()} --rm -v /data/:/data/ {fastp} fastp \
        -i {sample_dict["raw_fastq1"]} \
        -I {sample_dict["raw_fastq2"]} \
        -o {sample_dict["fastp1"]} \
        -O {sample_dict["fastp2"]} \
        -j {sample_dict["sample_fastq_dir"]}/{args.sample_id}.fastp.json \
        -h {sample_dict["sample_fastq_dir"]}/{args.sample_id}.fastp.html \
        -w {args.threads} \
        -x --detect_adapter_for_pe \
        --trim_poly_g'
    log_subprocess_output(f'fastp ({args.sample_id})',cmd)

######data alignment#####
def run_bwa():
    sample_dict["bwa_sam"] = f'{sample_dict["sample_preprocess_dir"]}/{args.sample_id}.bwa.sam'
    cmd = f'docker run -u {os.getuid()}:{os.getgid()} --rm -v /data/:/data/ {bwa} bwa mem \
        -t {args.threads} \
        -M {ref_fasta} \
        -o {sample_dict["bwa_sam"]} \
        {sample_dict["fastp1"]} \
        {sample_dict["fastp2"]}'
    log_subprocess_output(f'bwa mem ({args.sample_id})', cmd)


#######samtools #############
def run_samtools():
    cmd = f'docker run -u {os.getuid()}:{os.getgid()} --rm -v /data/:/data/ {samtools} samtools view \
        -bt {ref_fasta} \
        -o {sample_dict["sample_preprocess_dir"]}/{args.sample_id}.bwa.bam \
        {sample_dict["bwa_sam"]} ;'

    cmd += f'docker run -u {os.getuid()}:{os.getgid()} --rm -v /data/:/data/ {samtools} samtools sort \
        -o {sample_dict["sample_preprocess_dir"]}/{args.sample_id}.bwa.sorted.bam \
        {sample_dict["sample_preprocess_dir"]}/{args.sample_id}.bwa.bam;'

    cmd+=f'docker run -u {os.getuid()}:{os.getgid()} --rm -v /data/:/data/ {samtools} samtools index \
        {sample_dict["sample_preprocess_dir"]}/{args.sample_id}.bwa.sorted.bam'
    log_subprocess_output(f'run samtools ({args.sample_id})', cmd)

####################################################


###########################################
######2.CNV call (bicseq2) (.cnv)
###########################################
########make config
def make_project_config():
    bic_seq_dir = sample_dict["sample_bicseq2_dir"]
    bic_seq_config_dir = f'{bic_seq_dir}/config' # config 폴더 생성
    os.makedirs(f'{bic_seq_config_dir}', exist_ok=True)
    frag_size = 'FRAG_SIZE=200'
    read_length = 'READ_LENGTH=150'
    sample_config = f'{bic_seq_config_dir}/{args.sample_id}_config.sh'
    lines = [
    f'OUTD_BASE="{bic_seq_dir}"',
    'NORMD="$OUTD_BASE/norm"',
    'SEGD="$OUTD_BASE/segmentation"',
    'ANND="$OUTD_BASE/annotation"',
    'CHRLIST="/data/script/lims/lp_wgs/bicseq2/chromosomes.dat"',
    'REF_CHR="/data/script/lims/lp_wgs/HG19_CH"',
    'FA_CHR="${REF_CHR}/%s.fa"',
    'MAPD="/data/script/lims/lp_wgs/bicseq2"',
    'MER="hg19.CRC.100mer"',
    'MAP_CHR="$MAPD/$MER/$MER.%s.txt"',
    f'SEQD="{bic_seq_dir}/unique_mapping"',
    'SEQ_CHR="$SEQD/%s_%s.seq"',
    'SEQ_OUT="$SEQD/%s.seq"',
    f'{read_length}',
    f'{frag_size}',
    'BIN_SIZE=15000',
    'NORM_CHR="$NORMD/%s.%s.norm.bin"',
    'NORM_PDF="$NORMD/%s.GC.pdf"',
    'BICSEQ_NORM="/NBICseq-norm_v0.2.4/NBICseq-norm.pl"'
    ]
    with open(sample_config, 'w') as file:
        file.write('\n'.join(lines))
    file.close()
    print ("finish make_project_config")
    
###########get unique mapped read######
def get_unique_mapped_read(): 
    bic_seq_dir = sample_dict["sample_bicseq2_dir"]
    bicseq2_unique_dir = f'{bic_seq_dir}/unique_mapping'
    os.makedirs(f'{bicseq2_unique_dir}', exist_ok=True)
    cmd = ""
    for i in ['chr1','chr2','chr3','chr4','chr5','chr6','chr7','chr8','chr9','chr10',
              'chr11','chr12','chr13','chr14','chr15','chr16','chr17','chr18','chr19','chr20','chr21','chr22','chrX','chrY']:
        cmd += f'samtools view -@ 10 {sample_dict["sample_preprocess_dir"]}/{args.sample_id}.bwa.sorted.bam \
        {i} | perl /data/script/lims/lp_wgs/bicseq2/samtools-0.1.7a_getUnique-0.1.3/misc/samtools.pl unique - | cut -f 4 > {bicseq2_unique_dir}/{args.sample_id}_{i}.seq;'
    log_subprocess_output(f'get unique mapped read ({args.sample_id})', cmd)

# 2. Use BICseq2-norm to remove the biases in the data.

def run_bicseq2_norm():
    bic_seq_dir = sample_dict["sample_bicseq2_dir"]
    bic_seq_config_dir = f'{bic_seq_dir}/config'

    print(f'{bic_seq_config_dir}_config.sh')
    cmd = f'docker run -u {os.getuid()}:{os.getgid()} --rm -v /data/:/data/ {bicseq2} bash /BICSEQ2/src/run_norm.sh \
        {args.sample_id} {bic_seq_config_dir}/{args.sample_id}_config.sh'
    log_subprocess_output(f'run bicseq2 norm ({args.sample_id})', cmd)

# 3. Use BICseq2-seg to detect CNVs based on the normalized data.

def make_seg_config():
    bic_seq_dir = sample_dict["sample_bicseq2_dir"]
    bic_seq_config_dir = f'{bic_seq_dir}/config'
    norm_dir = f'{bic_seq_dir}/norm'
    chr_bin_list=[]
    for i in ['chr1','chr2','chr3','chr4','chr5','chr6','chr7','chr8','chr9','chr10','chr11','chr12','chr13','chr14','chr15','chr16','chr17','chr18','chr19','chr20','chr21','chr22','chrX','chrY']:
        chr_bin_list.append([i,f'{norm_dir}/{args.sample_id}.{i}.norm.bin'])
    seq_config = pd.DataFrame(chr_bin_list,columns= ['ChromName','binFileNorm'])
    seq_config.to_csv(f'{bic_seq_config_dir}/{args.sample_id}_seg_config.txt', sep='\t', index=False)
    print(f'{bic_seq_config_dir}/{args.sample_id}_seg_config.txt')

def run_bicseq2_seg():
    bic_seq_dir = sample_dict["sample_bicseq2_dir"]
    bic_seq_config_dir = f'{bic_seq_dir}/config'
    bic_seq_result_dir = f'{bic_seq_dir}/result'
    os.makedirs(f'{bic_seq_result_dir}', exist_ok=True)
    cmd =f'/data/script/lims/lp_wgs/bicseq2/NBICseq-seg_v0.7.2/NBICseq-seg.pl \
        --tmp {bic_seq_dir}/tmp\
        --fig {bic_seq_result_dir}/{args.sample_id}.png\
        --lambda=0.1\
        {bic_seq_config_dir}/{args.sample_id}_seg_config.txt \
        {bic_seq_result_dir}/{args.sample_id}.cnv'
    log_subprocess_output(f'run bicseq2 seg ({args.sample_id})', cmd)



################################################
######3. BICSEQ2 결과에 정보 추가 (y bin 값 사용 성별 정보/blacklist/backdata/cytoband overlap check, dgv frequency(정상 cnv) , gene 정보)
##############################################
##########filitering blacklist & backdata and add gene &dgv info 

###기준 통과한 CNV만 남기기
def bicseq_call_cut(cnv_call,dup_cf,del_cf,sex=None):
    cnv_call.rename(columns={'log2.copyRatio':'log2_copyRatio'}, inplace = True)
    auto_chr_cutoff = (~cnv_call['chrom'].isin(['chrX', 'chrY'])) & ((cnv_call['log2_copyRatio'] >= dup_cf) | (cnv_call['log2_copyRatio'] <= del_cf))
    x_chr_cutoff = (cnv_call['chrom']=='chrX') & (
    (sex == 'M') & ((cnv_call['log2_copyRatio'] >= (dup_cf + (-1))) | (cnv_call['log2_copyRatio'] <= (del_cf + (-1)))) |
    (sex != 'M') & ((cnv_call['log2_copyRatio'] >= dup_cf) | (cnv_call['log2_copyRatio'] <= del_cf))
    )
    y_chr_cutoff = (cnv_call['chrom']== 'chrY') & (
    (sex == 'M') & ((cnv_call['log2_copyRatio'] >= (dup_cf + (-1))) | (cnv_call['log2_copyRatio'] <= (del_cf + (-1)))) 
    ) #XY
    cnv_cut = cnv_call[auto_chr_cutoff|x_chr_cutoff|y_chr_cutoff].copy()
    if len(cnv_cut)==0:
        return (cnv_cut)
    cnv_cut['pos'] = cnv_cut['start'].astype(str) + '-' + cnv_cut['end'].astype(str)
    cnv_cut['segment_size'] = cnv_cut['end'] - cnv_cut['start']
    cnv_cut['chrom'] = cnv_cut['chrom'].str.replace("chr", "")#.replace({'X': 23, 'Y': 24}).astype(int)
    cnv_cut['cnv_type'] = ['dup' if x >= dup_cf else 'del' for x in cnv_cut['log2_copyRatio']] 
    if sex =='M':
        cnv_cut.loc[(cnv_cut['chrom'].isin(['X', 'Y'])) & (cnv_cut['log2_copyRatio'] >= (dup_cf + (-1))),'cnv_type'] ='dup'
        cnv_cut.loc[(cnv_cut['chrom'].isin(['X', 'Y'])) & (cnv_cut['log2_copyRatio'] <= (del_cf + (-1))),'cnv_type'] ='del'
    cnv_cut_fin = cnv_cut[cnv_cut.columns.difference(['binNum','observed','expeceted'])]
    return(cnv_cut_fin)

### NORMAL(=BACKDATA)/ call된 CNV 간 overlap check 
def calculate_overlap_cnv(start1, end1, start2, end2, cnv_type1, cnv_type2):
    if cnv_type1 != cnv_type2:
        overlap = 0
    else:
        overlap = max(0, min(end1, end2) - max(start1, start2))
    return overlap

def calculate_overlap(start1, end1, start2, end2):
    overlap = max(0, min(end1, end2) - max(start1, start2))
    return overlap

def overlap_percent(df1,df2):
    if len(df2)==0:
        return(df2)
    overlaps = []
    cyto_info =[]
    for row1 in df1.itertuples(index=False):
        for row2 in df2.itertuples(index=False):
            if row1.Chr != row2.chrom :
                continue
            overlap = calculate_overlap(row1.Start, row1.End, row2.start, row2.end)
            if overlap != 0:
                overlaps.append(row2)
                cyto_info.append(row1.Cytoband)
    overlaps_df = pd.DataFrame(overlaps)
    overlaps_df['cyto'] = cyto_info
    col_list = overlaps_df.columns.difference(['cyto']).to_list()
    cyto_df = overlaps_df.groupby(col_list)['cyto'].apply(','.join).reset_index()
    return(cyto_df)

##blacklist overlap 추가
def blacklist_overlap_info(blacklist,cnv_list):
    if len(cnv_list)==0:
        return(cnv_list)
    no_overlaps = []
    for bl in blacklist.itertuples(index=False):
        for cnv in cnv_list.itertuples(index=False):
            if (bl.chrom == cnv.chrom):
                overlap = calculate_overlap(bl.start, bl.end, cnv.start, cnv.end)
                if (overlap !=0):
                    no_overlaps.append({'chrom':cnv.chrom, 'start':cnv.start, 'end':cnv.end, 'log2_copyRatio':cnv.log2_copyRatio, 'pos':cnv.pos,
                        'cnv_type':cnv.cnv_type, 'segment_size':cnv.segment_size, 'blacklist_ovelap_p':(overlap/cnv.segment_size)*100,'blacklist_overlap':overlap,'blacklist_info':bl.region_info})
                elif (overlap ==0):
                    no_overlaps.append({'chrom':cnv.chrom, 'start':cnv.start, 'end':cnv.end, 'log2_copyRatio':cnv.log2_copyRatio, 'pos':cnv.pos,
                        'cnv_type':cnv.cnv_type, 'segment_size':cnv.segment_size, 'blacklist_ovelap_p':0,'blacklist_overlap':0,'blacklist_info':'-'})
    no_overlaps_df = pd.DataFrame(no_overlaps)
    no_overlaps_df.drop_duplicates(inplace=True)
    return(no_overlaps_df)

##normal overlap 추가
def normal_overlap_info(normal,cnv_list):
    if len(cnv_list)==0:
        return(cnv_list)
    no_overlaps = []
    for nor in normal.itertuples(index=False):
        for cnv in cnv_list.itertuples(index=False):
            if cnv.chrom not in normal['chrom']:
                no_overlaps.append({'chrom':cnv.chrom, 'start':cnv.start, 'end':cnv.end, 'log2_copyRatio':cnv.log2_copyRatio, 'pos':cnv.pos,
                        'cnv_type':cnv.cnv_type, 'segment_size':cnv.segment_size,'blacklist_info':cnv.blacklist_info,'blacklist_overlap_p':cnv.blacklist_ovelap_p,
                        'blacklist_overlap':cnv.blacklist_overlap,'backdata_info':'-','backdata_percent':'-','backdata_intersect_p':0,
                        'backdata_dgv_info':'-'})
            if (nor.chrom == cnv.chrom):
                overlap = calculate_overlap_cnv(nor.start, nor.end, cnv.start, cnv.end, nor.cnv_type, cnv.cnv_type)
                if (overlap !=0):
                    no_overlaps.append({'chrom':cnv.chrom, 'start':cnv.start, 'end':cnv.end, 'log2_copyRatio':cnv.log2_copyRatio, 'pos':cnv.pos,
                        'cnv_type':cnv.cnv_type, 'segment_size':cnv.segment_size,'blacklist_info':cnv.blacklist_info, 'blacklist_overlap_p':cnv.blacklist_ovelap_p,
                        'blacklist_overlap':cnv.blacklist_overlap,'backdata_info':nor.backdata_info,'backdata_percent':nor.normal_percent,
                        'backdata_intersect_p':(overlap/cnv.segment_size)*100,
                        'backdata_dgv_info':nor.dgv_info_all})
                elif (overlap ==0):
                    no_overlaps.append({'chrom':cnv.chrom, 'start':cnv.start, 'end':cnv.end, 'log2_copyRatio':cnv.log2_copyRatio, 'pos':cnv.pos, 
                        'cnv_type':cnv.cnv_type, 'segment_size':cnv.segment_size,'blacklist_info':cnv.blacklist_info,'blacklist_overlap_p':cnv.blacklist_ovelap_p,
                        'blacklist_overlap':cnv.blacklist_overlap,'backdata_info':'-','backdata_percent':'-','backdata_intersect_p':0,
                        'backdata_dgv_info':'-'})
    no_overlaps_df = pd.DataFrame(no_overlaps)
    no_overlaps_df.drop_duplicates(inplace=True)
    return(no_overlaps_df)



###calling cnv - overlap dgv frequency 추가

def max_frequency_check(df1, df2):
    df1['backdata_index'] = df1.index
    # Merge the two dataframes based on 'chrom' and 'variant_sub_type'
    merged_df = pd.merge(df1, df2, left_on=['chrom', 'cnv_type'], right_on=['chrom', 'variant_sub_type'], how='left')
    merged_df = merged_df[(~merged_df['dgv_start'].isnull().values) | (~merged_df['dgv_end'].isnull().values)]
    # Calculate overlap for each row
    merged_df['intersect'] = merged_df.apply(lambda row: calculate_overlap(row['start'], row['end'], row['dgv_start'], row['dgv_end']), axis=1)
    

    # Calculate intersect percent for each row
    merged_df['dgv_intersect_percent'] = (merged_df['intersect'] / merged_df['segment_size'])

    # Filter rows with intersect_percent >= 80
    filtered_df = merged_df.loc[merged_df['dgv_intersect_percent'] >= 0.8]
    # Group by 'backdata_info' and keep the row with the maximum 'frequency'
    high_intersect_backdata_info = filtered_df.loc[filtered_df.groupby('backdata_index')['frequency'].idxmax()]
    high_intersect_backdata_info['dgv_cutoff_pass'] = 'Y'
    # add backdata row with intersect percent <80
    row_intersect_backdata_info = merged_df[~merged_df['backdata_index'].isin(high_intersect_backdata_info['backdata_index'].tolist())]
    row_intersect_backdata_fin = row_intersect_backdata_info.loc[row_intersect_backdata_info.groupby('backdata_index')['dgv_intersect_percent'].idxmax()]
    row_intersect_backdata_fin.loc[row_intersect_backdata_fin['intersect']==0,'frequency']=0
    row_intersect_backdata_fin['dgv_cutoff_pass'] ='N'

    result_df = pd.concat([high_intersect_backdata_info,row_intersect_backdata_fin], sort=False)
    result_df['dgv_info'] = 'chr' +result_df['chrom']+': '+ result_df['dgv_start'].astype(str)+'-' +result_df['dgv_end'].astype(str)
    result_df.sort_values(by=['backdata_index'],inplace=True)


    return result_df

###gene 정보 추가 
def gene_overlap_info(gene_info, cnv_list):
    if cnv_list.empty:
        return cnv_list

    no_overlaps = []

    for cnv in cnv_list.itertuples(index=False):
        matching_chrom = gene_info[gene_info['chrom'] == cnv.chrom]

        if matching_chrom.empty:
            no_overlaps.append(create_overlap_entry(cnv))
        else:
            for gene in matching_chrom.itertuples(index=False):
                overlap = calculate_overlap(gene.start, gene.end, cnv.start, cnv.end)

                if overlap != 0:
                    no_overlaps.append(create_overlap_entry(cnv, gene, overlap))
                else:
                    no_overlaps.append(create_overlap_entry(cnv))

    no_overlaps_df = pd.DataFrame(no_overlaps)
    no_overlaps_df.drop_duplicates(inplace=True)
    return no_overlaps_df

def create_overlap_entry(cnv, gene=None, overlap=0):
    entry = {
        'chrom': cnv.chrom,
        'start': cnv.start,
        'end': cnv.end,
        'log2_copyRatio': cnv.log2_copyRatio,
        'pos': cnv.pos,
        'cnv_type': cnv.cnv_type,
        'segment_size': cnv.segment_size,
        'cyto': cnv.cyto,
        'blacklist_info': cnv.blacklist_info,
        'blacklist_overlap_p': cnv.blacklist_overlap_p,
        'blacklist_overlap': cnv.blacklist_overlap,
        'backdata_info': cnv.backdata_info,
        'backdata_percent': cnv.backdata_percent,
        'backdata_intersect_p': cnv.backdata_intersect_p,
        'backdata_dgv_info': cnv.backdata_dgv_info,
        'Gene_name':'-',
        'Gene_info':'-'
    }

    if gene:
        entry['Gene_name'] = gene.Gene_name
        entry['Gene_info'] = gene.Gene_info

    return entry

#######연결되는 CNV 합치기 
def can_merge_rows(row1, row2):
    return (
        row2['chrom'] == row1['chrom']
        and row2['start'] == row1['end'] + 1
        and row2['cnv_type'] == row1['cnv_type']
    )

def merge_rows(row1, row2):
    new_row = row1.copy()
    new_row['end'] = row2['end']
    new_row['log2_copyRatio'] = ((row1['end'] - row1['start']) / (row2['end'] - row1['start'])) * row1['log2_copyRatio'] + \
                                ((row2['end'] - row2['start']) / (row2['end'] - row1['start'])) * row2['log2_copyRatio']
    new_row['segment_size'] = row1['segment_size'] +  row2['segment_size'] + 1
    return new_row

def continuous_row_merge(bicseq_result):
    bicseq_fin = bicseq_result.sort_values(['chrom','start'])
    bicseq_fin.reset_index(drop=True,inplace=True)
    for i in range(1,len(bicseq_fin)):
        prev_row = bicseq_fin.loc[i-1]
        curr_row = bicseq_fin.loc[i]
        if can_merge_rows(prev_row, curr_row):
            merged_row = merge_rows(prev_row, curr_row)
            bicseq_fin.drop(index=[i-1,i], inplace=True)
            bicseq_fin.loc[i]= merged_row
        else:
            i += 1
    bicseq_fin2 = bicseq_fin.sort_index()
#    bicseq_cut = bicseq_fin[(bicseq_fin['cnv_type'].isin(['dup','del']))]
    return bicseq_fin2
###############
def sex_chromosome_call_merge(cnv_fin):
    cnv_result = cnv_fin.copy()
    col_len =len(cnv_result.columns)
    x = cnv_result[cnv_result['chrom']=='X']
    y = cnv_result[cnv_result['chrom']=='Y']
    if (len(x)>0) & (len(x['cnv_type'].unique())==1) & (x['segment_size'].sum()/155270560>=0.8):
        cnv_result.drop(x.index,inplace = True)
        cnv_result.loc[x.index[0]] = ['X',1,155270560,x['log2_copyRatio'].mean(),'1-155270560',x['cnv_type'].unique()[0],155270560,'all'] + ['-']*(col_len-8)
    if (len(y)>0) & (len(y['cnv_type'].unique())==1) & (y['segment_size'].sum()/59373566>=0.8): 
        cnv_result.drop(y.index,inplace = True)
        cnv_result.loc[y.index[0]] = ['Y',1,59373566,y['log2_copyRatio'].mean(),'1-59373566',y['cnv_type'].unique()[0],59373566,'all'] + ['-']*(col_len-8)
    cnv_result.reset_index(drop=True,inplace=True)
    return cnv_result

dgv_hg19 = pd.read_csv('/data/analysis/project/ref/DGV_hg19_essential_info.csv')

def blacklist_n_back_filitering():
    table_all_dir = f'{sample_dict["sample_result_dir"]}/1.Table'
    os.makedirs(f'{table_all_dir}', exist_ok=True)

    sample_id = args.sample_id
    sample_id_origin = args.sample_id
    #sample sex info
    #back data 
    bd_dgv = pd.read_excel(f'/data/analysis/project/Result/230727_A01980_0012_BH5LNGDRX3/backdata_based_0_8_add_dgv.xlsx')
    bd_dgv['chrom'] = bd_dgv['chrom'].astype(str)
    #black list 
    blacklist = pd.read_csv('/data/home/heewon/ref/hg19-blacklist.v2.bed', index_col = 0) 
    blacklist['chrom'] = blacklist['chr'].str.replace("chr", "")
    blacklist['region_info'] = 'chr'+blacklist['chrom']+':'+blacklist['start'].astype(str)+'-'+blacklist['end'].astype(str)+' ('+blacklist['info']+')'
    
    #whole chromosome list
    chrom_list2 = list(map(str,[*range(1,23)]+['X','Y']))
    #cut off list
    black_cutoff= 35
    back_cutoff = 15

    #dgv hg19 info
    
    #cytoband
    cytoband = pd.read_csv("/data/home/heewon/ref/cytoBand.csv")
    cytoband['Chr'] = cytoband['Chr'].astype(str)

    #gene info 
    gene_info = pd.read_csv('/data/analysis/project/ref/bio_mart_source_info_exist_gene.csv')
    gene_info['chrom'] = gene_info['chrom'].astype(str)

    #Y bin 값 확인하여 성별 정보 체크 
    bic_seq_dir = sample_dict["sample_bicseq2_dir"]

    bicseq2_bin = f'{bic_seq_dir}/norm/{sample_id}'
    y_bin = pd.read_csv( f'{bicseq2_bin}.chrY.norm.bin',sep='\t')
    log2_copy_ratio = np.array([-8]*(len(y_bin)))  # -inf convert to -8
    log2_copy_ratio = log2_copy_ratio.astype(float)
    non_zero_mask = y_bin['obs'] != 0  # Create a mask for non-zero values in 'obs' column
    log2_copy_ratio[non_zero_mask] = np.log2(np.divide(y_bin['obs'][non_zero_mask], y_bin['expected'][non_zero_mask]))
    sample_sex_info = 'M' if np.median(log2_copy_ratio)>-2 else 'F'

    ###최종파일 .cnv 저장된 dir
    bic_seq_dir = sample_dict["sample_bicseq2_dir"]
    bic_seq_result_dir = f'{bic_seq_dir}/result'

    bic_result = pd.read_csv(f'{bic_seq_result_dir}/{sample_id}.cnv',sep='\t')
    
    ##DEL, DUP CUT OFF (현재는 2카피 기준 2.6 카피 이상 DUP/ 1.4 이하 DEL)
    sample_cnv = bicseq_call_cut(bic_result,math.log2(2.6/2),math.log2(1.4/2),sex=sample_sex_info)
    dup_cnv = sample_cnv[sample_cnv['cnv_type']=='dup']
    del_cnv = sample_cnv[sample_cnv['cnv_type']=='del']

    ##한 base 차이로 이어지는 CNV는 연결 
    dup_fin = continuous_row_merge(dup_cnv)
    del_fin = continuous_row_merge(del_cnv)
    merged_cnv = pd.concat([dup_fin,del_fin])
    
    ##SIZE 기준 CUT OFF = 200000(200KB) -> 해당 크기 이상 CNV는 테이블로 확인 )
    sample_cnv_size_cut = merged_cnv[merged_cnv['segment_size']>=200000]

    no_bl_cut = blacklist_overlap_info(blacklist,sample_cnv_size_cut)
    bl_with_non_overlap = no_bl_cut[(no_bl_cut[['chrom', 'start', 'end', 'log2_copyRatio', 'pos', 'segment_size']].duplicated(keep=False)) & (no_bl_cut['blacklist_info']=='-')].index
    no_bl_cut.drop(bl_with_non_overlap,inplace=True)
    no_bl_cut_result = no_bl_cut.groupby(['chrom', 'start', 'end', 'log2_copyRatio', 'pos', 'segment_size', 'cnv_type'])\
        .agg({'blacklist_info': ','.join, 'blacklist_overlap': sum,'blacklist_ovelap_p':sum})\
        .reset_index()
    normal_bl_info_add = normal_overlap_info(bd_dgv,no_bl_cut_result)
    normal_with_non_overlap = normal_bl_info_add[(normal_bl_info_add[['chrom', 'start', 'end', 'log2_copyRatio', 'pos', 'segment_size']].duplicated(keep=False)) & (normal_bl_info_add['backdata_percent']==0)].index
    normal_bl_info_add.drop(normal_with_non_overlap,inplace=True)
    normal_bl_info_add['backdata_percent'] =normal_bl_info_add['backdata_percent'].astype(str)
    normal_bl_info_add['chrom'] = pd.Categorical(normal_bl_info_add['chrom'], 
    categories=chrom_list2, 
    ordered=True)
    n_fin  = overlap_percent(cytoband,normal_bl_info_add)
    n_fin['segment_size_KB'] = n_fin['segment_size'].apply(lambda x: f"{int(x / 1000)}KB" if x % 1000 == 0 else f"{x / 1000:.1f}KB")
    n_fin_add_bd = n_fin.groupby(['chrom', 'start', 'end', 'log2_copyRatio', 'pos', 'segment_size','segment_size_KB', 'cnv_type','cyto','blacklist_info','blacklist_overlap','blacklist_overlap_p'])\
        .agg({'backdata_info':','.join,'backdata_dgv_info':','.join,'backdata_percent': ','.join, 'backdata_intersect_p': sum})\
        .reset_index()
    n_fin_add_bd['backdata_dgv_info'] = n_fin_add_bd['backdata_dgv_info'].str.replace("-,", "")
    n_fin_add_bd['backdata_percent'] = n_fin_add_bd['backdata_percent'].str.replace("0,", "")
    n_fin_add_bd['backdata_info'] = n_fin_add_bd['backdata_info'].str.replace("-,", "")

    n_fin_add_bd['chrom'] = pd.Categorical(n_fin_add_bd['chrom'], 
    categories=chrom_list2, 
    ordered=True)
    n_fin_add_bd.sort_values(by=['chrom','start'],inplace=True)

    n_fin_add_bd_cut = n_fin_add_bd

    if n_fin_add_bd_cut.empty: 
        empty_col_list = n_fin_add_bd_cut.columns.to_list()
        empty_df = pd.DataFrame([pd.Series(['-']*len(empty_col_list),index=empty_col_list)])
        empty_df.to_excel(f'{table_all_dir}/{sample_id_origin}.xlsx',index=None) 
    else:
        g_fin = gene_overlap_info(gene_info, n_fin_add_bd_cut)
        g_fin_col_list = g_fin.columns.difference(['Gene_name','Gene_info']).to_list()
        g_fin_df = g_fin.groupby(g_fin_col_list)\
        .agg({'Gene_name': ','.join, 'Gene_info': ','.join}).reset_index()
        fin_df_col_list = ['chrom','start','end','log2_copyRatio','pos','cnv_type','segment_size','cyto','blacklist_info','blacklist_overlap_p','blacklist_overlap',
           'backdata_info','backdata_percent','backdata_intersect_p','backdata_dgv_info','Gene_name','Gene_info']
        g_fin_df2 = g_fin_df[fin_df_col_list]

        g_fin_df2['Gene_name'] = g_fin_df2['Gene_name'].str.replace("-,", "")
        g_fin_df2['Gene_info'] = g_fin_df2['Gene_info'].str.replace("-,", "")
        dgv_info_add_fin_df = max_frequency_check(g_fin_df2,dgv_hg19)
        dgv_info_add_fin_df.drop_duplicates(inplace=True)
        dgv_info_add_fin_df['frequency'] = dgv_info_add_fin_df['frequency'].astype(str)
        dgv_info_add_fin_df2 = dgv_info_add_fin_df.groupby(fin_df_col_list)\
        .agg({'dgv_info': ','.join, 'frequency': ','.join,'dgv_cutoff_pass':','.join}).reset_index()
        dgv_info_add_fin_df2['chrom'] = pd.Categorical(dgv_info_add_fin_df2['chrom'], 
        categories=chrom_list2, 
        ordered=True)
        dgv_info_add_fin_df2 = sex_chromosome_call_merge(dgv_info_add_fin_df2)
        dgv_info_add_fin_df2.sort_values(by=['chrom','start'],inplace=True)
        dgv_info_add_fin_df2.rename(columns={'blacklist_overlap_p':'blacklist_intersect_p','blacklist_overlap':'blacklist_intersect_bp','backdata_percent':'backdata_include_sample_p'},inplace=True)
        dgv_info_add_fin_df2.loc[dgv_info_add_fin_df2['backdata_include_sample_p']=='1','backdata_include_sample_p'] = '83/83'
        cnv_count = len(dgv_info_add_fin_df2)
        add_id_sex = pd.DataFrame({'sample_id':[sample_id]*cnv_count,'sex':[sample_sex_info]*cnv_count})
        fin_excel = pd.concat([add_id_sex,dgv_info_add_fin_df2],axis=1)
        
        fin_excel['chrom'] = fin_excel['chrom'].astype(str)
        fin_excel['chrom'] = pd.Categorical(fin_excel['chrom'], 
        categories=chrom_list2, 
        ordered=True)
        fin_excel.sort_values(by=['chrom','start'],inplace=True)
        fin_excel.reset_index(drop=True,inplace=True)
        fin_excel.insert(0,'cnv_id','CNV' + (fin_excel.index +1).astype(str))
        fin_excel.to_excel(f'{table_all_dir}/{sample_id_origin}.xlsx',index=None) 

    print(f'{sample_id_origin} sample filitering & add info finish')

    return fin_excel


###########################################
######CUT OFF 통과 CNV 색깔 추가 ###########
###########################################

def add_color():
    sample_id = args.sample_id
    table_all_dir = f'{sample_dict["sample_result_dir"]}/1.Table'
    result_excel =f"{table_all_dir}/{sample_id}.xlsx"
    wb  = load_workbook(result_excel)
    df_value_exist = pd.read_excel(result_excel)
    df_value_exist['backdata_intersect_p'] = df_value_exist['backdata_intersect_p'].replace('-', 0)
    df_value_exist['blacklist_intersect_p'] = df_value_exist['blacklist_intersect_p'].replace('-', 0)
    # Assuming df_value_exist is your DataFrame
    del_index = df_value_exist[(df_value_exist['backdata_intersect_p'] <= 15) & 
                            (df_value_exist['blacklist_intersect_p'] <= 33.693) & 
                            (df_value_exist['cnv_type'] == 'del')].index.tolist()

    dup_index = df_value_exist[(df_value_exist['backdata_intersect_p'] <= 15) & 
                            (df_value_exist['blacklist_intersect_p'] <= 33.693) & 
                            (df_value_exist['cnv_type'] == 'dup')].index.tolist()
    

    if len(del_index)>0:
        for row in del_index:
            for col in range(df_value_exist.shape[1]):
                wb.active.cell(row+2,col+1).fill = PatternFill(start_color="F6B2B5", end_color="F6B2B5", fill_type="solid")
    
    if len(dup_index)>0:
        for row in dup_index:
            for col in range(df_value_exist.shape[1]):
                wb.active.cell(row+2,col+1).fill = PatternFill(start_color="8FCBF0", end_color="8FCBF0", fill_type="solid")  # change hex code to change color
    
    wb.save(f"{table_all_dir}/{sample_id}.xlsx")



#######bin result by tool ###########
def bicseq_plotly_format(bic_fin,number=None):
    bin_list=[]
    if number is None:
        chrom = [*range(1,23)]+['X','Y']
        for i in chrom:
            bin_info=pd.read_csv(f'{bic_fin}.chr{i}.norm.bin',\
                      sep='\t',engine='python')
            bin_info['CHR'] = i
            bin_list.append(bin_info)
        bin_fin = pd.concat(bin_list,axis=0)
    else:
        bin_fin=pd.read_csv(f'{bic_fin}.chr{number}.norm.bin',\
                        sep='\t',engine='python')
        bin_fin['CHR'] = number
    bin_fin.loc[bin_fin['CHR']=='X','CHR'] = 23
    bin_fin.loc[bin_fin['CHR']=='Y','CHR'] = 24
    log2_copy_ratio = np.array([-8]*len(bin_fin))  # -inf convert to -8
    log2_copy_ratio = log2_copy_ratio.astype(float)
    non_zero_mask = bin_fin['obs'] != 0  # Create a mask for non-zero values in 'obs' column
    log2_copy_ratio[non_zero_mask] = np.log2(np.divide(bin_fin['obs'][non_zero_mask], bin_fin['expected'][non_zero_mask]))
    # Assign the calculated values to the DataFrame
    bin_fin['log2_copyRatio'] = log2_copy_ratio
    bin_fin = bin_fin.dropna()
    bin_fin['bin_mid']= ((bin_fin['start']+bin_fin['end'])//2).astype(int)
    bin_fin['pos']=bin_fin['start'].astype(str) + '-'+ bin_fin['end'].astype(str)
    return bin_fin

def blacklist_figure(blacklist,number=None):
    data=blacklist[blacklist['chr']==f'chr{number}']
    delta = data['end']-data['start']
    pos = data['start'].astype(str)+'-'+data['end'].astype(str)
    if type(number) is str:
        title="chromosome %s" % number
    else:
        title="chromosome %i" % number

    layout = go.Layout(
           title=title,
           xaxis={'showgrid': False},
                  yaxis={'title': 'Blacklist'},
           hovermode='closest')
    hover_text = 'Position: '+ pos \
             + '<br>Information: '+ data['info']
    data_to_plot = []
    data_to_plot.append((go.Bar(
         y=[0]*len(delta),
         x=delta,
         base=data['start'],
         name='Blacklist',
         orientation='h',
         width=1,
         hovertemplate=hover_text,
         marker=dict(color='rgba(0, 0, 0, 1)')
                )
            )
         )
    return go.Figure(data=data_to_plot, layout=layout, layout_yaxis_range=[0,0])
##simulation data answer

#

def bicseq_call_chrom(cnv_call,dup_cf,del_cf,sex=None,number=None):
    cnv_call.rename(columns={'log2.copyRatio':'log2_copyRatio'}, inplace = True)
    if number is None:
        chr_cutoff = (~cnv_call['chrom'].isin(['chrX','chrY'])) &((cnv_call['log2_copyRatio']>=dup_cf)|(cnv_call['log2_copyRatio']<= del_cf))
        #chrY_cutoff = (cnv_call['chrom'].isin(['chrX','chrY'])) &((cnv_call['log2_copyRatio']>=cutoff)|(cnv_call['log2_copyRatio']<= -cutoff))
        if sex=='M':
            sex_chr_cutoff = (cnv_call['chrom'].isin(['chrX','chrY'])) &((cnv_call['log2_copyRatio']>=dup_cf+(-1))|(cnv_call['log2_copyRatio']<= del_cf+(-1)))
        else:
            sex_chr_cutoff = (cnv_call['chrom'].isin(['chrX','chrY'])) &((cnv_call['log2_copyRatio']>=dup_cf)|(cnv_call['log2_copyRatio']<= del_cf))
        cnv_cut = cnv_call[chr_cutoff | sex_chr_cutoff].copy()
    else:
        if sex=='M':
            if number=='X'|number=='Y':
                cnv_cut = cnv_call[(cnv_call['chrom']==f'chr{number}') &((cnv_call['log2_copyRatio']>=dup_cf+(-1))|(cnv_call['log2_copyRatio']<= del_cf+(-1)))].copy()
            else:
                cnv_cut = cnv_call[(cnv_call['chrom']==f'chr{number}') &((cnv_call['log2_copyRatio']>=dup_cf)|(cnv_call['log2_copyRatio']<= del_cf))].copy()
        else:
            cnv_cut = cnv_call[(cnv_call['chrom']==f'chr{number}') &((cnv_call['log2_copyRatio']>=dup_cf)|(cnv_call['log2_copyRatio']<= del_cf))].copy()

    if len(cnv_cut)==0:
        return (cnv_cut)
    cnv_cut['pos'] = cnv_cut['start'].astype(str)+'-'+cnv_cut['end'].astype(str)
    cnv_cut['size'] = cnv_cut['end'] - cnv_cut['start'] +1
    cnv_cut['tool']='bic_seq2'
    cnv_cut['chrom']=cnv_cut['chrom'].str.replace("chr","")
    cnv_cut.loc[cnv_cut['chrom']=='X','chrom']= 23
    cnv_cut.loc[cnv_cut['chrom']=='Y','chrom']= 24
    cnv_cut['chrom'] = cnv_cut['chrom'].astype(int)
    return(cnv_cut)



def calculate_overlap(start1, end1, start2, end2):
    overlap = max(0, min(end1, end2) - max(start1, start2))
    return overlap


def bin_figure(bin_result,blacklist,sex=None,call=None,number=None,normal=None):
    if 'X' in call['chrom'].unique().tolist():
        call['chrom'] = call['chrom'].replace({'X': 23})
    if 'Y' in call['chrom'].unique().tolist():
        call['chrom'] = call['chrom'].replace({'Y': 24})
    call['chrom'] = call['chrom'].astype(int)
    normal['position'] = normal['start'].astype(str)+'-'+normal['end'].astype(str)
    data_to_plot = []
    bl_plot = []
    cl_plot = []
    nor_plot =[]
    bl=blacklist.copy()
    point_size=5
    yrange=[-8,8]
    if number is None:
        if normal is not None:
            nor = normal.copy()
            nor['delta'] = nor['end']-nor['start']
            nor['pos'] = nor['start'].astype(str)+'-'+nor['end'].astype(str)
        cl=call.copy()
        cl['delta'] = cl['end']-cl['start']
        cl['pos'] = cl['start'].astype(str)+'-'+cl['end'].astype(str)
        data=bin_result.copy()
        data['pos'] = data['start'].astype(str)+'-'+data['end'].astype(str)
        data['log2_copyRatio_cut'] = data['log2_copyRatio']
        data.loc[data['log2_copyRatio']<yrange[0],'log2_copyRatio_cut'] = yrange[0]
        data.loc[data['log2_copyRatio']>yrange[1],'log2_copyRatio_cut'] = yrange[1]
        bl.loc[bl['chr']== 'chrX','chr']= 'chr23'
        bl.loc[bl['chr']== 'chrY','chr']= 'chr24'

        bl['CHR'] = bl['chr'].str.extract(r'chr([0-9]*)').astype(int)
        lastbase = 0
        ticks = []
        ticksLabels = []
        col=None
        nchr = len(data['CHR'].unique())
        ticksLabels = data['CHR'].unique().tolist()
        ticksLabels = ['X' if value == 23 else 'Y' if value == 24 else value for value in ticksLabels]
        for i in data['CHR'].unique():
            if i == 1:
                data.loc[data['CHR'] == i, 'bin_mid_chr'] = \
                data.loc[data['CHR'] == i, 'bin_mid'].values

                bl.loc[bl['CHR'] == i, 'start_chr'] = \
                bl.loc[bl['CHR'] == i, 'start'].values
                bl.loc[bl['CHR'] == i, 'end_chr'] = \
                bl.loc[bl['CHR'] == i, 'end'].values

                cl.loc[cl['chrom'] == i, 'start_chr'] = \
                cl.loc[cl['chrom'] == i, 'start'].values
                cl.loc[cl['chrom'] == i, 'end_chr'] = \
                cl.loc[cl['chrom'] == i, 'end'].values

                nor.loc[nor['chrom'] == i, 'start_chr'] = \
                nor.loc[nor['chrom'] == i, 'start'].values
                nor.loc[nor['chrom'] == i, 'end_chr'] = \
                nor.loc[nor['chrom'] == i, 'end'].values
            else:
                prevbin = data.loc[data['CHR'] == i - 1,'bin_mid']
                lastbase = lastbase + prevbin.iat[-1]
                data.loc[data['CHR'] == i, 'bin_mid_chr'] = \
                data.loc[data['CHR'] == i, 'bin_mid'].values + lastbase

                bl.loc[bl['CHR'] == i, 'start_chr'] = \
                        bl.loc[bl['CHR'] == i, 'start'].values + lastbase
                bl.loc[bl['CHR'] == i, 'end_chr'] = \
                bl.loc[bl['CHR'] == i, 'end'].values + lastbase

                cl.loc[cl['chrom'] == i, 'start_chr'] = \
                        cl.loc[cl['chrom'] == i, 'start'].values + lastbase
                cl.loc[cl['chrom'] == i, 'end_chr'] = \
                cl.loc[cl['chrom'] == i, 'end'].values + lastbase
                nor.loc[nor['chrom'] == i, 'start_chr'] = \
                        nor.loc[nor['chrom'] == i, 'start'].values + lastbase
                nor.loc[nor['chrom'] == i, 'end_chr'] = \
                nor.loc[nor['chrom'] == i, 'end'].values + lastbase

            tmin = min(data.loc[data['CHR'] == i, 'bin_mid_chr'])
            tmax = max(data.loc[data['CHR'] == i, 'bin_mid_chr'])
            ticks.append(int((tmin + tmax) / 2.) + 1)

        icol = 0
        if col is None:
            col = ['black' if np.mod(i, 2)
                   else 'grey' for i in range(nchr)]    
        for i in data['CHR'].unique():
                tmp = data[data['CHR'] == i]
                tmp_cl = cl[cl['chrom'] == i]
                tmp_nor = nor[nor['chrom'] == i]
                if len(tmp_nor)>0:
                    nor_dup = tmp_nor[tmp_nor["log2_copyRatio"]>0]
                    nor_dele = tmp_nor[tmp_nor["log2_copyRatio"]<0]
                else:
                    nor_dup = []
                    nor_dele = []
                if len(tmp_cl)>0:
                    dup = tmp_cl[tmp_cl["log2_copyRatio"]>0]
                    dele = tmp_cl[tmp_cl["log2_copyRatio"]<0]
                else:
                    dup = []
                    dele = []
                if i==23:
                    ii = 'X'
                    names = "Chr %s" % ii
                    if sex=='M':
                        nor_dup = tmp_nor[tmp_nor["log2_copyRatio"]>-1]
                        nor_dele = tmp_nor[tmp_nor["log2_copyRatio"]<-1]
                        dup = tmp_cl[tmp_cl["log2_copyRatio"]>-1]
                        dele = tmp_cl[tmp_cl["log2_copyRatio"]<-1]
                elif i==24:
                    ii='Y'
                    names = "Chr %s" % ii
                    if sex=='M':
                        nor_dup = tmp_nor[tmp_nor["log2_copyRatio"]>-1]
                        nor_dele = tmp_nor[tmp_nor["log2_copyRatio"]<-1]
                        dup = tmp_cl[tmp_cl["log2_copyRatio"]>-1]
                        dele = tmp_cl[tmp_cl["log2_copyRatio"]<-1]
                else:
                    ii=i
                    names = "Chr %i" % ii
                col_list = [col[icol]]*len(tmp)
                if len(dup)>0:
                    for row_n, row1 in tmp.iterrows():
                        for j, row2 in dup.iterrows():
                        # Compare the start and end columns
                            if row1['start'] >= row2['start'] and row1['end'] <= row2['end']:
                            # Change the color in the color list
                                col_list[row_n] = 'rgba(4, 12, 246, 0.6)'
                if len(dele)>0:
                    for row_n, row1 in tmp.iterrows():
                        for j, row2 in dele.iterrows():
                        # Compare the start and end columns
                            if row1['start'] >= row2['start'] and row1['end'] <= row2['end']:
                            # Change the color in the color list
                                col_list[row_n] = 'rgba(246, 4, 83, 0.6)'
                #chromo = tmp['chr'].unique()  # Get chromosome name
                data_to_plot.append(
                    go.Scattergl(
                        x=tmp['bin_mid_chr'],
                        y=tmp['log2_copyRatio_cut'],
                        name=names,
                        mode="markers",
                        marker={
                            'color': col_list,
                            'size': point_size,
                            'opacity': 0.2
                        },
                        text='Position: ' + tmp['pos'].astype(str)\
                            + '<br>log2ratio: ' + tmp['log2_copyRatio'].astype(str)
                    )
                )
                ##blacklist
                tmp_bl = bl[bl['CHR'] == i]
                delta = tmp_bl['end_chr']-tmp_bl['start_chr']
                pos = tmp_bl['start'].astype(str)+'-'+tmp_bl['end'].astype(str)
                bl_plot.append((go.Bar(
                    y=[0]*len(delta),
                    x=delta,
                    base=tmp_bl['start_chr'],
                    showlegend=False,
                    name=names,
                    orientation='h',
                    width=3,
                    hovertemplate='Position: '+ pos \
                       + '<br>Information: '+ tmp_bl['info'],
                    marker=dict(color='rgba(0, 0, 0, 1)')
                        )
                    )
                )
                ##normal answer
                if len(tmp_nor)>0:
                    if sex=='M':
                        nor_dup = tmp_nor[(((tmp_nor['chrom'].isin([23,24])) & (tmp_nor["log2_copyRatio"]>-1)) | ((~tmp_nor['chrom'].isin([23,24])) & (tmp_nor["log2_copyRatio"]>0)))]
                        nor_dele = tmp_nor[(((tmp_nor['chrom'].isin([23,24])) & (tmp_nor["log2_copyRatio"]<-1)) | ((~tmp_nor['chrom'].isin([23,24])) & (tmp_nor["log2_copyRatio"]<0)))]
                    else:
                        nor_dup = tmp_nor[tmp_nor["log2_copyRatio"]>0]
                        nor_dele = tmp_nor[tmp_nor["log2_copyRatio"]<0]
                    if len(nor_dup)>0:
                        nor_plot.append((go.Bar(
                            y=[0]*len(nor_dup['log2_copyRatio']),
                            x=nor_dup['delta'],
                            base=nor_dup['start_chr'],
                            name='duplication',
                            showlegend=False,
                            orientation='h',
                            width=3,
                            hovertemplate='<br>Position: '+ nor_dup['position']\
                                + '<br>Size: '+ nor_dup['delta'].astype(str)\
                                + '<br>include_sample_p: ' + nor_dup['normal_percent'],
                            marker=dict(color='rgba(4, 12, 246, 0.2)')
                                )
                            )   
                        )
                    if len(nor_dele)>0:
                        nor_plot.append((go.Bar(
                        y=[0]*len(nor_dele['log2_copyRatio']),
                        x=nor_dele['delta'],
                        base=nor_dele['start_chr'],
                        name='deletion',
                        showlegend=False,
                        orientation='h',
                        width=3,
                            hovertemplate='<br>Position: '+ nor_dele['position']\
                                + '<br>Size: '+ nor_dele['delta'].astype(str)\
                                + '<br>include_sample_p: ' + nor_dele['normal_percent'],
                        marker=dict(color='rgba(246, 4, 83, 0.2)')
                               )
                           )
                        )
                ##calling answer
                if len(tmp_cl)>0:
                    if sex=='M':
                        dup = tmp_cl[(((tmp_cl['chrom'].isin([23,24])) & (tmp_cl["log2_copyRatio"]>-1)) | ((~tmp_cl['chrom'].isin([23,24])) & (tmp_cl["log2_copyRatio"]>0)))]
                        dele = tmp_cl[(((tmp_cl['chrom'].isin([23,24])) & (tmp_cl["log2_copyRatio"]<-1)) | ((~tmp_cl['chrom'].isin([23,24])) & (tmp_cl["log2_copyRatio"]<0)))]
                    else:
                        dup = tmp_cl[tmp_cl["log2_copyRatio"]>0]
                        dele = tmp_cl[tmp_cl["log2_copyRatio"]<0]
                    if len(dup)>0:
                        cl_plot.append((go.Bar(
                            y=[0]*len(dup['log2_copyRatio']),
                            x=dup['delta'],
                            base=dup['start_chr'],
                            name='duplication',
                            showlegend=False,
                            orientation='h',
                            width=3,
                            hovertemplate='Position: '+ dup['pos'] \
                                + '<br>Cytoband: '+ dup['cyto']\
                                + '<br>Size: '+ dup['delta'].astype(str)\
                                + '<br>log2ratio: ' + dup['log2_copyRatio'].astype(str)\
                                +'<br>backdata_intersect_p: '+ dup['backdata_intersect_p'].astype(str)\
                                +'<br>blacklist_intersect_p: '+dup['blacklist_intersect_p'].astype(str),
                            marker=dict(color='rgba(4, 12, 246, 0.2)')
                                )
                            )   
                        )
                    if len(dele)>0:
                        cl_plot.append((go.Bar(
                        y=[0]*len(dele['log2_copyRatio']),
                        x=dele['delta'],
                        base=dele['start_chr'],
                        name='deletion',
                        showlegend=False,
                        orientation='h',
                        width=3,
                        hovertemplate='Position: '+ dele['pos']\
                            + '<br>Cytoband: '+ dele['cyto']\
                            + '<br>Size: '+ dele['delta'].astype(str)\
                            + '<br>log2ratio: ' + dele['log2_copyRatio'].astype(str)\
                            +'<br>backdata_intersect_p: '+ dele['backdata_intersect_p'].astype(str)\
                            +'<br>blacklist_intersect_p: '+dele['blacklist_intersect_p'].astype(str),
                        marker=dict(color='rgba(246, 4, 83, 0.2)')
                               )
                           )
                        )
                icol = icol + 1
    else:
        if number == 'X':
            number = 23
        if number == 'Y':
            number = 24
        data=bin_result.copy()[bin_result['CHR']==number]
        data['log2_copyRatio_cut'] = data['log2_copyRatio']
        data.loc[data['log2_copyRatio']<yrange[0],'log2_copyRatio_cut'] = yrange[0]
        data.loc[data['log2_copyRatio']>yrange[1],'log2_copyRatio_cut'] = yrange[1]
        layout = go.Layout(
             xaxis={
                    'title': "Chromosome %i" % number,
                    'showgrid': False,
                    'tickmode': "array",
                    'ticks': "outside"
                },
                hovermode='closest'
                )
        data_to_plot.append(
            go.Scattergl(
                x=data['bin_mid'],
                y=data['log2_copyRatio_cut'],
                name="%s" % data['TOOL'].unique()[0],
                mode="markers",
                marker={
                        'color': 'black',
                        'size': point_size,
                        'opacity': 0.2
                    },
                    text='Position: ' + data['pos'].astype(str)\
                        + '<br>log2ratio: ' + data['log2_copyRatio'].astype(str)
                )
            )
        ##calling result
        if call is not None:
            cl=call
            cl['delta'] = cl['end']-cl['start']
            cl['pos'] = cl['start'].astype(str)+'-'+cl['end'].astype(str)
            tmp_cl = cl[cl['chrom'] == number]
            if len(tmp_cl)>0:
                if sex=='M':
                    dup = tmp_cl[(((tmp_cl['chrom'].isin([23,24])) & (tmp_cl["log2_copyRatio"]>-1)) | ((~tmp_cl['chrom'].isin([23,24])) & (tmp_cl["log2_copyRatio"]>0)))]
                    dele = tmp_cl[(((tmp_cl['chrom'].isin([23,24])) & (tmp_cl["log2_copyRatio"]<-1)) | ((~tmp_cl['chrom'].isin([23,24])) & (tmp_cl["log2_copyRatio"]<0)))]
                else:
                    dup = tmp_cl[tmp_cl["log2_copyRatio"]>0]
                    dele = tmp_cl[tmp_cl["log2_copyRatio"]<0]
                if len(dup)>0:
                    data_to_plot.append((go.Bar(
                                    y=[0]*len(dup['log2_copyRatio']),
                                    x=dup['delta'],
                                    base=dup['start'],
                                    name='duplication',
                                    showlegend=False,
                                    orientation='h',
                                    width=3,
                                    hovertemplate='Position: '+ dup['pos'] \
                                        + '<br>Cytoband: '+ dup['cyto']\
                                        + '<br>Size: '+ dup['delta'].astype(str)\
                                        + '<br>log2ratio: ' + dup['log2_copyRatio'].astype(str)\
                                        +'<br>backdata_intersect_p'+ dup['backdata_intersect_p'].astype(str)\
                                        +'<br>blacklist_intersect_p'+dup['blacklist_intersect_p'].astype(str),
                                    marker=dict(color='rgba(4, 12, 246, 0.6)')
                                        )
                                    )   
                                )
                if len(dele)>0:
                    data_to_plot.append((go.Bar(
                                y=[0]*len(dele['log2_copyRatio']),
                                x=dele['delta'],
                                base=dele['start'],
                                name='deletion',
                                orientation='h',
                                showlegend=False,
                                width=3,
                                hovertemplate='Position: '+ dele['pos'] \
                                    + '<br>Cytoband: '+ dele['cyto']\
                                    + '<br>Size: '+ dele['delta'].astype(str)\
                                    + '<br>log2ratio: ' + dele['log2_copyRatio'].astype(str)\
                                    +'<br>backdata_intersect_p'+ dele['backdata_intersect_p'].astype(str)\
                                    +'<br>blacklist_intersect_p'+dele['blacklist_intersect_p'].astype(str),
                                marker=dict(color='rgba(246, 4, 83, 0.6)')
                                       )
                                   )
                                )
    if number is None:
        fig = make_subplots(rows=4, cols=1,
                shared_xaxes=True,
                shared_yaxes=True,
                row_heights = [15/600,15/600,15/600,555/600],
                vertical_spacing=0.02)
        fig.update_layout(template="seaborn",width=1800)
        #normal plot
        if len(nor_plot)>0:
            for i in range(0,len(nor_plot)):
                fig.add_trace(nor_plot[i],row=1, col=1)
            #,showticklabels=False
        fig.update_yaxes(range=[0,0],showticklabels=False,showgrid=False,zeroline=False,row=1, col=1)
        fig.update_xaxes(showgrid=False,zeroline=False,row=1, col=1)
        #call plot
        if len(cl_plot)>0:
            for i in range(0,len(cl_plot)):
                fig.add_trace(cl_plot[i],row=2, col=1)
            #,showticklabels=False
        fig.update_yaxes(range=[0,0],showticklabels=False,showgrid=False,zeroline=False,row=2, col=1)
        fig.update_xaxes(showgrid=False,zeroline=False,row=2, col=1)
        #blacklist plot
        for i in range(0,len(bl_plot)):
            fig.add_trace(bl_plot[i],row=3, col=1)
        fig.update_yaxes(range=[0,0],showticklabels=False,showgrid=False,zeroline=False,row=3, col=1)
        fig.update_xaxes(showgrid=False,zeroline=False,row=3, col=1)
        
        #sample result bin plot 
        for i in range(0,len(data_to_plot)):
            fig.add_trace(data_to_plot[i],row=4, col=1)
        fig.update_xaxes(title_text="Chromosomes",showgrid = False,tickmode="array",tickvals=ticks,ticktext=ticksLabels,ticks="outside",range=[min(data_to_plot[0].x),max(data_to_plot[23].x)], row=4, col=1)
        fig.update_yaxes(title_text="Log2 copy ratio",range=[-8,8],showgrid=False,zeroline=True, row=4, col=1)

        return fig
    else:
        return data_to_plot

def whole_bin_plot():
    sample_id = args.sample_id
    bic_seq_dir = sample_dict["sample_bicseq2_dir"]
    result_dir = sample_dict["sample_result_dir"]
    ######필요파일
    blacklist = pd.read_csv('/data/home/heewon/ref/hg19-blacklist.v2.bed', index_col = 0) 
    #######
    table_all_dir = f'{result_dir}/1.Table'
    whole_bin_dir = f'{result_dir}/2.Plot/whole_bin'


    os.makedirs(whole_bin_dir, exist_ok=True)

    bicseq2_bin = f'{bic_seq_dir}/norm/{sample_id}'
    
    bicseq2_call = pd.read_excel(f'{table_all_dir}/{sample_id}.xlsx')

    y_bin = pd.read_csv( f'{bicseq2_bin}.chrY.norm.bin',sep='\t')
    log2_copy_ratio = np.array([-8]*(len(y_bin)))  # -inf convert to -8
    log2_copy_ratio = log2_copy_ratio.astype(float)
    non_zero_mask = y_bin['obs'] != 0  # Create a mask for non-zero values in 'obs' column
    log2_copy_ratio[non_zero_mask] = np.log2(np.divide(y_bin['obs'][non_zero_mask], y_bin['expected'][non_zero_mask]))
    sample_sex_info = 'M' if np.median(log2_copy_ratio)>-2 else 'F'
    ###back data with dgv info ########
    bd_dgv = pd.read_excel(f'/data/analysis/project/Result/230727_A01980_0012_BH5LNGDRX3/backdata_based_0_8_add_dgv.xlsx')
    ################
    
    bicseq2_bin_plot=bin_figure(bicseq_plotly_format(bicseq2_bin),blacklist,sex=sample_sex_info,call=bicseq2_call,normal=bd_dgv)
    #bicseq2_bin_plot.write_html(f'{plotly_result_dir_together}/{sample_id}_lambda_0_1.html')
    #pip install kaleido 깔려있어야 write_image 정상 작동 
    bicseq2_bin_plot.write_image(f'{whole_bin_dir}/{sample_id}.png')
    print(f'whole bin plot {sample_id}')

#CNV 한개에 대한 plot 코드 


    

def fin_bar_plot():
    sample_id = args.sample_id
    bic_seq_dir = sample_dict["sample_bicseq2_dir"]
    result_dir = sample_dict["sample_result_dir"]
    bic_bin_fin = f'{bic_seq_dir}/norm/{sample_id}'
    table_all_dir = f'{result_dir}/1.Table'

    all_chr_bin = bicseq_plotly_format(bic_bin_fin)
    #whole_bin_dir = f'{result_dir}/2.Plot/whole_bin'
    each_cnv_plot_dir = f'{result_dir}/2.Plot/cnv_bin' #전처리 된 결과만 저장
    os.makedirs(f'{each_cnv_plot_dir}', exist_ok=True)

    blacklist = pd.read_csv('/data/home/heewon/ref/hg19-blacklist.v2.bed', index_col = 0) 
    normal = pd.read_excel(f'/data/analysis/project/230412_lp_wgs_heewon/Result/230727_A01980_0012_BH5LNGDRX3/backdata_based_0_8_add_dgv.xlsx')
    result_table = pd.read_excel(f"{table_all_dir}/{sample_id}.xlsx")

    all_chr_bin['chr']=all_chr_bin['CHR'].astype(str)
    all_chr_bin.loc[all_chr_bin['chr'] =='23','chr'] = 'X'
    all_chr_bin.loc[all_chr_bin['chr'] =='24','chr'] = 'Y'
    #cnv 별 plot 
    ##tmp -> 각 플랏에 필요한 데이터만 임시 저장 
    for j in range(0,len(result_table)):
        ###cnv info 
        cnv = result_table.iloc[j,:]
        if cnv["blacklist_intersect_p"]=='-':
            cnv["blacklist_intersect_p"]=0
        if cnv["backdata_intersect_p"]=='-':
            cnv["backdata_intersect_p"]=0
        cnv_info = f'▲chr{cnv["chrom"]}:{cnv["pos"]}({cnv["cyto"]})<br>blacklist intersect p: {round(float(cnv["blacklist_intersect_p"]),3)} / backdata intersect p: {round(float(cnv["backdata_intersect_p"]),3)}'
        cnv_start_point,cnv_end_point = cnv['start'],cnv['end'] #400000
    
        ######플랏에 필요한 bin 정보 
        tmp_bin = all_chr_bin[(all_chr_bin['chr']==str(cnv['chrom']))& (all_chr_bin['start'] >=cnv_start_point-1000000) & (all_chr_bin['end'] <=cnv_end_point+1000000)].reset_index(drop=True)

        ###bin plot color
        cnv_color_info = tmp_bin[(tmp_bin['start']>=cnv_start_point) & (tmp_bin['end']<=cnv_end_point)].index.tolist()
        cnv_color ='DUP' if cnv['cnv_type']=='dup' else 'DELE'
        color_list = [cnv_color if i in cnv_color_info else 'NORMAL' for i in range(len(tmp_bin))]
        tmp_bin['TYPE'] = color_list 
        ######tmp blacklist 
        blacklist['chr_str']=blacklist['chr'].str.replace('chr', '')
        bl_chr = blacklist[blacklist['chr_str'] == cnv['chrom']]
        tmp_bl = bl_chr[(bl_chr['start'] >=cnv_start_point-10000000) & (bl_chr['start'] <=cnv_end_point+10000000)].reset_index(drop=True)

        normal['chr_str'] = normal['chrom'].astype(str)
        nor_chr = normal[normal['chr_str'] == cnv['chrom']]
        tmp_nor = nor_chr[(nor_chr['start'] >=cnv_start_point-10000000) & (nor_chr['start'] <=cnv_end_point+10000000)].reset_index(drop=True)

    
        ###bar plot data 
        bar_start=[]
        bar_end = []
        bar_y=[]
        bar_color = []

        if len(tmp_bl)>0:
            bar_start.extend(tmp_bl['start'].values.tolist())
            bar_end.extend(tmp_bl['end'].values.tolist())
            bar_y.extend(['blacklist']*len(tmp_bl))
            bar_color.extend(['black']*len(tmp_bl))
        else:
            bar_start.append(cnv['start'])
            bar_end.append(cnv['end'])
            bar_y.extend(['blacklist'])
            bar_color.extend(['white'])

        #CNV plot 
        bar_start.append(cnv['start'])
        bar_end.append(cnv['end'])
        bar_y.extend(['cnv'])
        cnv_color = ['blue' if cnv['cnv_type'] =='dup' else 'red']
        bar_color.extend(cnv_color)

        if len(tmp_nor)>0:
            bar_start.extend(tmp_nor['start'].values.tolist())
            bar_end.extend(tmp_nor['end'].values.tolist())
            bar_y.extend(['nor']*len(tmp_nor))
            nor_color = ['blue' if tmp_nor['cnv_type'][i] == 'dup' else 'red' for i in range(len(tmp_nor))]
            bar_color.extend(nor_color)
        else:
            bar_start.append(cnv['start'])
            bar_end.append(cnv['end'])
            bar_y.extend(['normal'])
            bar_color.extend(['white'])
        
        #cnv info 
        cnv_info = f'▲chr{cnv["chrom"]}:{cnv["pos"]}({cnv["cyto"]})<br>blacklist intersect p: {round(float(cnv["blacklist_intersect_p"]),3)} / backdata intersect p: {round(float(cnv["backdata_intersect_p"]),3)}'

        sns.set_style("white")
        # Sample data for the horizontal bar plot

        bar_data = {
            'x_start': bar_start,
            'x_end': bar_end,
            'y': bar_y,
            'color': bar_color
        }


        df = pd.DataFrame(bar_data)

        # Convert color names to RGBA values
        df['color_rgba'] = df['color'].apply(lambda x: sns.xkcd_rgb[x])

        ##NORMAL DUP DELE
        # Create a JointGrid with scatter plot
        chr_all = sns.JointGrid(data=tmp_bin, x="bin_mid", y="log2_copyRatio")
        palette ={"NORMAL": "black", "DELE": "red", "DUP": "blue"}
        #black blue red
        chr_all.plot_joint(sns.scatterplot,hue=tmp_bin['TYPE'],palette=palette)
        plt.ylim((-4, 4))
        plt.xlim((cnv_start_point-400000,cnv_end_point+400000))

        # Add horizontal bar plot to the marginal x-axis
        # Add horizontal bar plot to the marginal x-axis
        bar_height = 0.3
        bar_padding = 0  # Adjust this value to control the gap between bars

        #bar_padding = 0.02  # Adjust this value to control the gap between bars

        # Add horizontal bar plot to the marginal x-axis
        #bar_height = 0.3

        # Add horizontal bar plot to the marginal x-axis
        for i, y_value in enumerate(df['y'].unique()):
            for _, row in df[df['y'] == y_value].iterrows():
                bar_width = row['x_end'] - row['x_start']
                chr_all.ax_marg_x.barh(i - 0.1, width=bar_width, left=row['x_start'],
                                       height=0.3, color=row['color_rgba'], alpha=0.7)

        chr_all.set_axis_labels(xlabel=cnv_info, ylabel='log2_copyRatio', fontsize=14)
        chr_all.ax_joint.set_xticks([])
        chr_all.ax_marg_x.set_xticks([])
        chr_all.ax_marg_y.set_xticks([])
        chr_all.ax_marg_x.set_xlim((cnv_start_point - 400000, cnv_end_point + 400000))
        cnv_name = cnv['cnv_id']
        cnv_size = round(cnv['segment_size']/1000,2)
        cnv_info = f'chr{cnv["chrom"]}:{cnv["pos"]} ({cnv_size} KB)'
        
        
        chr_all.fig.suptitle(cnv_info, y=1.02, x = 0.45, fontsize=16)
        
        plt.draw()
        #chr_all.fig.set_size_inches(5, 5)
        chr_all.savefig(f'{each_cnv_plot_dir}/{sample_id}_{cnv_name}.png')
        plt.close(chr_all.fig)
        print(f'{cnv_name}_plot_fin')
    
    ########전체그림/ 전체 CNV 취합 
    # Number of columns in the grid
    num_columns = 3
    num_plots = len(result_table) 
    num_rows = (num_plots + num_columns - 1) // num_columns

    f, axarr = plt.subplots(num_rows + 1, num_columns, figsize=(25, num_rows*8))

    ax_all_p = f.add_axes([0, 0.5, 1, 0.75])  # [left, bottom, width, height]
    #ax_all_p.imshow(mpimg.imread('all_p.png'))
    whole_bin_dir = f'{result_dir}/2.Plot/whole_bin'
    all_chr = f'{whole_bin_dir}/{sample_id}.png'
    ax_all_p.imshow(mpimg.imread(all_chr))
    ax_all_p.axis('off')  # Turn off the axis for the 'all_p.png' image

    for k in range(0,num_plots):
        row = k // num_columns + 1
        col = k % num_columns    
        axarr[row, col].imshow(mpimg.imread(f'{each_cnv_plot_dir}/{sample_id}_CNV{k+1}.png'))

    # Load and display other images in the second row (replace 'g0.png' with your actual image file names)
    #axarr[1, 0].imshow(mpimg.imread('chr_all.png'))
    #axarr[1, 1].imshow(mpimg.imread('chr_all.png'))
    #axarr[1, 2].imshow(mpimg.imread('chr_all.png'))

    # Turn off x and y axis for all subplots
    [ax.set_axis_off() for ax in axarr.ravel()]

    # Save the entire grid as a PDF

    plot_result_dir = f'{result_dir}/2.Plot/fin_pdf'
    os.makedirs(plot_result_dir, exist_ok=True)
    
    plt.savefig(f'{plot_result_dir}/{sample_id}.pdf')

    # Show the plot (optional)



################################################################
###########cnv call visualization###############################
################################################################



def fin_table_to_chr_plot():
    sample_id = args.sample_id
    table_all_dir = f'{sample_dict["sample_result_dir"]}/1.Table'
    result_excel =f'{table_all_dir}/{sample_id}.xlsx'
    cnv_table = pd.read_excel(result_excel)
    cnv_table[['blacklist_intersect_p', 'backdata_intersect_p']] = cnv_table[['blacklist_intersect_p', 'backdata_intersect_p']].replace('-', 0)
    result_dir = sample_dict["sample_result_dir"]
    chr_plot_dir = f'{result_dir}/2.Plot/chromosome' #전처리 된 결과만 저장
    os.makedirs(f'{chr_plot_dir}', exist_ok=True)

    cnv_cutoff= cnv_table[(cnv_table['backdata_intersect_p'] <= 15) & 
                            (cnv_table['blacklist_intersect_p'] <= 33.693) & 
                            (cnv_table['segment_size'] >= 400000)].copy()

    cnv_cutoff.loc[:,'color'] = cnv_cutoff['cnv_type'].apply(lambda x: '#77F' if x == 'dup' else '#f97d83')

    annotations_list = []

    for _, row in cnv_cutoff.iterrows():
        annotation = (
            f"{{"
            f"chr: '{str(row['chrom'])}', "
            f"start: {int(row['start'])}, "
            f"stop: {int(row['end'])}, "
            f"color: '{row['color']}', "
            f"shape: 'circle'"
            f"}}"
        )
        annotations_list.append(annotation)

    annotations_code = ', '.join(annotations_list)

    javascript_code = f"""
    <head>
        <script src="https://cdn.jsdelivr.net/npm/ideogram@1.45.1-beta/dist/js/ideogram.min.js"></script>
      </head>
      <body>
          <div class="small-ideogram">
      <script>
    
        var config = {{
          container: '.small-ideogram',
          orientation: 'vertical',
          organism: 'human',
          resolution: 550,
          chrHeight: 300,
          rows: 2,
          rotatable: false,
          annotations: [{annotations_code}]
        }};
        var ideogram = new Ideogram(config);
      </script>
    </body>
    </html>
    """

    # Save to HTML file
    output_file_path = f'{chr_plot_dir}/{args.sample_id}.html'
    with open(output_file_path, 'w') as html_file:
        html_file.write(javascript_code)

    print(f"HTML content saved to {output_file_path}")


def finish_summary():
    #spend time as a min
    end_time = time.time()
    spend_time = (end_time - start_time)/60    

    logging.info(f'{"-"*20}{"Summary":^50}{"-"*20}')
    logging.info(f'{"sample_id":<15}: {args.sample_id}')
    logging.info(f'{"run_id":<15}: {args.run_id}')
    logging.info(f'{"fastq1":<15}: {sample_dict["raw_fastq1"]}')
    logging.info(f'{"fastq2":<15}: {sample_dict["raw_fastq2"]}')
    logging.info(f'{"sample_fastq_dir":<15}: {sample_dict["sample_fastq_dir"]}')
    logging.info(f'{"sample_preprocess_dir":<15}: {sample_dict["sample_preprocess_dir"]}')
    logging.info(f'{"sample_result_dir":<15}: {sample_dict["sample_result_dir"]}')
    logging.info(f'{"spend time":<15}: {spend_time:.2f} min')
    logging.info(f'{"pipeline ver":<15}: 1.0.0')
    logging.info(f'{"-"*20}{"Pipeline Finished":^50}{"-"*20}')

from_email = 'totoaje11@gmail.com'
to_email = 'totoaje11@gmail.com'
def log_subprocess_output(step, cmd):
    global error_occurred
    logging.info(f'{"processing":<15}: {step:<30}')
    logging.debug(f'$ {cmd}')
    #subprocess 이후 진행 상황을 file_handler에 저장
    process = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
    while True:
        line = process.stdout.readline()
        if not line :
            break
        logging.debug('> %s', line.decode('utf-8').strip())
    
    process.wait()

    if process.returncode != 0 and not error_occurred:  # subprocess가 성공하지 못했을 때 ERROR로 기록
        error_message = f'Error occurred in subprocess: {step}'
        logging.error(f'{error_message}')

        with open(f'{sample_dict["sample_dir"]}/err.txt', 'w') as test_file:
            test_file.write('An error occurred in the logging.')
        mail_subject = f"[lp_wgs ERROR] {args.run_id} : {args.sample_id} "
        mail_contents = f"\t !!!!!!!!!!!!!! Please check the error !!!!!!!!!!!!!! \n \t \n {error_message} \n \t \n Run ID : {args.run_id} \n \t Sample ID : {args.sample_id}  \n \t \n \t  !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!! "
        send_email_command = f'echo "{mail_contents}" | /usr/bin/mail -s "{mail_subject}" -r {from_email} {to_email}' # from, to
        subprocess.run(send_email_command, shell=True)
        error_occurred = True

    return error_occurred

sample_dict["sample_dir"] = f'{args.output_dir}/{args.run_id}/{args.sample_id}'
os.makedirs(sample_dict["sample_dir"], exist_ok=True)

def send_completion_email():
    global error_occurred
    if not error_occurred:  # 에러가 발생하지 않은 경우 분석 완료 메일 보내기
        with open(f'{sample_dict["sample_dir"]}/complete.txt', 'w') as test_file:
            test_file.write('complete analysis.')
        mail_subject = f"[lp_wgs complete] {args.run_id} : {args.sample_id} "
        mail_contents = f" \n \t Run ID : {args.run_id} \n \t Sample ID : {args.sample_id}  \n \t "
        send_email_command = f'echo "{mail_contents}" | /usr/bin/mail -s "{mail_subject}" -r {from_email} {to_email}' # from, to
        subprocess.run(send_email_command, shell=True)

sample_dict["sample_dir"] = f'{args.output_dir}/{args.run_id}/{args.sample_id}'
os.makedirs(sample_dict["sample_dir"], exist_ok=True)


###########################################################################

def main():
    global error_occurred
    error_occurred = False
    
    init_env_check()
    
    #preprocess
    run_fastp()
    run_bwa()
    run_samtools()

    #cnv call
    make_project_config()
    get_unique_mapped_read()
    run_bicseq2_norm()
    make_seg_config()
    run_bicseq2_seg()

    #cnv filitering 
    blacklist_n_back_filitering() 

    #cutoff pass CNV 색깔 추가
    add_color()


    #visualization 
    whole_bin_plot()
    fin_bar_plot()
    fin_table_to_chr_plot()

    #anaylsis fin
    finish_summary()
    send_completion_email()

if __name__ == "__main__":
    main()


###########all_sample_info_xlsx
